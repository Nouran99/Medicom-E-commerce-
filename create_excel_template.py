#!/usr/bin/env python3
"""
Create Excel template for Medicum Egypt product import
with sample data and instructions
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
import json
from datetime import datetime

# Sample product data for Medicum Egypt
products_data = [
    {
        "product_code": "MED-001-PARA",
        "name_en": "Paracetamol 500mg Tablets",
        "name_ar": "باراسيتامول 500 مجم أقراص",
        "description_en": "Pain relief and fever reduction tablets. Effective for headaches, muscle aches, and mild to moderate pain.",
        "description_ar": "أقراص لتخفيف الألم وخفض الحرارة. فعال للصداع وآلام العضلات والألم الخفيف إلى المتوسط.",
        "category": "pain-relief",
        "subcategory": "oral-medications",
        "brand": "PharmaCare",
        "manufacturer": "Egyptian Pharma Co.",
        "country_of_origin": "Egypt",
        "price_per_unit": 15.50,
        "currency": "EGP",
        "unit": "box",
        "unit_size": "20 tablets",
        "stock_quantity": 500,
        "min_order_quantity": 1,
        "max_order_quantity": 10,
        "sku": "PARA-500-20T",
        "barcode": "6221234567891",
        "hsn_code": "30049099",
        "requires_prescription": False,
        "is_controlled": False,
        "active_ingredient": "Paracetamol 500mg",
        "dosage_form": "Tablet",
        "side_effects": "Rare: nausea, allergic reactions. Overdose can cause liver damage.",
        "contraindications": "Severe liver disease, allergy to paracetamol",
        "storage_conditions": "Store below 30°C in a dry place",
        "expiry_date": "2026-12-31",
        "batch_number": "B2024-001",
        "seller_code": "SELLER-001",
        "delivery_method": "standard",
        "delivery_days_min": 1,
        "delivery_days_max": 3,
        "delivery_fee": 30.00,
        "free_delivery_threshold": 200.00,
        "return_policy": "7 days return for unopened packages",
        "warranty_period": "N/A",
        "discount_percentage": 10,
        "tax_percentage": 14,
        "product_images": json.dumps([
            "https://example.com/images/paracetamol-1.jpg",
            "https://example.com/images/paracetamol-2.jpg"
        ]),
        "specifications": json.dumps({
            "form": "Round tablet",
            "color": "White",
            "package": "Blister pack"
        }),
        "tags": "pain-relief,fever,headache,OTC",
        "meta_title": "Buy Paracetamol 500mg Online | Medicum Egypt",
        "meta_description": "Paracetamol 500mg tablets for effective pain relief. Fast delivery across Egypt.",
        "is_featured": True,
        "is_active": True,
        "weight_grams": 50,
        "dimensions_cm": "10x8x3"
    },
    {
        "product_code": "MED-002-AMOX",
        "name_en": "Amoxicillin 500mg Capsules",
        "name_ar": "أموكسيسيلين 500 مجم كبسولات",
        "description_en": "Broad-spectrum antibiotic for bacterial infections. Treats respiratory, urinary tract, and skin infections.",
        "description_ar": "مضاد حيوي واسع الطيف للعدوى البكتيرية. يعالج التهابات الجهاز التنفسي والمسالك البولية والجلد.",
        "category": "antibiotics",
        "subcategory": "oral-medications",
        "brand": "MediCore",
        "manufacturer": "Cairo Pharmaceuticals",
        "country_of_origin": "Egypt",
        "price_per_unit": 45.00,
        "currency": "EGP",
        "unit": "box",
        "unit_size": "14 capsules",
        "stock_quantity": 300,
        "min_order_quantity": 1,
        "max_order_quantity": 5,
        "sku": "AMOX-500-14C",
        "barcode": "6221234567907",
        "hsn_code": "30041090",
        "requires_prescription": True,
        "is_controlled": False,
        "active_ingredient": "Amoxicillin trihydrate 500mg",
        "dosage_form": "Capsule",
        "side_effects": "Common: diarrhea, nausea. Rare: allergic reactions, rash.",
        "contraindications": "Penicillin allergy, severe kidney disease",
        "storage_conditions": "Store below 25°C in a dry place",
        "expiry_date": "2026-06-30",
        "batch_number": "B2024-002",
        "seller_code": "SELLER-001",
        "delivery_method": "express",
        "delivery_days_min": 1,
        "delivery_days_max": 2,
        "delivery_fee": 50.00,
        "free_delivery_threshold": 300.00,
        "return_policy": "No returns for prescription medications",
        "warranty_period": "N/A",
        "discount_percentage": 0,
        "tax_percentage": 14,
        "product_images": json.dumps([
            "https://example.com/images/amoxicillin-1.jpg",
            "https://example.com/images/amoxicillin-2.jpg",
            "https://example.com/images/amoxicillin-3.jpg"
        ]),
        "specifications": json.dumps({
            "form": "Hard gelatin capsule",
            "color": "Red/Yellow",
            "package": "Blister pack"
        }),
        "tags": "antibiotic,prescription,infection,amoxicillin",
        "meta_title": "Amoxicillin 500mg Capsules | Prescription Required",
        "meta_description": "Amoxicillin 500mg antibiotic capsules. Prescription required. Fast delivery.",
        "is_featured": True,
        "is_active": True,
        "weight_grams": 40,
        "dimensions_cm": "12x8x2"
    },
    {
        "product_code": "MED-003-VITD",
        "name_en": "Vitamin D3 5000 IU Softgels",
        "name_ar": "فيتامين د3 5000 وحدة دولية كبسولات",
        "description_en": "High-strength Vitamin D3 supplement for bone health and immune support.",
        "description_ar": "مكمل فيتامين د3 عالي القوة لصحة العظام ودعم المناعة.",
        "category": "vitamins",
        "subcategory": "supplements",
        "brand": "VitaHealth",
        "manufacturer": "Egyptian Nutrition Labs",
        "country_of_origin": "Egypt",
        "price_per_unit": 85.00,
        "currency": "EGP",
        "unit": "bottle",
        "unit_size": "60 softgels",
        "stock_quantity": 200,
        "min_order_quantity": 1,
        "max_order_quantity": 6,
        "sku": "VITD3-5000-60",
        "barcode": "6221234567914",
        "hsn_code": "21069099",
        "requires_prescription": False,
        "is_controlled": False,
        "active_ingredient": "Cholecalciferol (Vitamin D3) 5000 IU",
        "dosage_form": "Softgel",
        "side_effects": "Rare: hypercalcemia with excessive doses",
        "contraindications": "Hypercalcemia, kidney stones",
        "storage_conditions": "Store in a cool, dry place",
        "expiry_date": "2027-03-31",
        "batch_number": "B2024-003",
        "seller_code": "SELLER-002",
        "delivery_method": "standard",
        "delivery_days_min": 2,
        "delivery_days_max": 4,
        "delivery_fee": 30.00,
        "free_delivery_threshold": 250.00,
        "return_policy": "30 days return for unopened bottles",
        "warranty_period": "N/A",
        "discount_percentage": 15,
        "tax_percentage": 14,
        "product_images": json.dumps([
            "https://example.com/images/vitd3-1.jpg",
            "https://example.com/images/vitd3-2.jpg"
        ]),
        "specifications": json.dumps({
            "form": "Softgel",
            "color": "Golden yellow",
            "package": "Plastic bottle"
        }),
        "tags": "vitamin,supplement,bone-health,immune-support,OTC",
        "meta_title": "Vitamin D3 5000 IU - 60 Softgels | Medicum Egypt",
        "meta_description": "High-strength Vitamin D3 supplement for optimal health. 60 softgels.",
        "is_featured": False,
        "is_active": True,
        "weight_grams": 120,
        "dimensions_cm": "6x6x10"
    },
    {
        "product_code": "MED-004-METF",
        "name_en": "Metformin 500mg Tablets",
        "name_ar": "ميتفورمين 500 مجم أقراص",
        "description_en": "Oral medication for type 2 diabetes management. Helps control blood sugar levels.",
        "description_ar": "دواء فموي لإدارة مرض السكري من النوع 2. يساعد في السيطرة على مستويات السكر في الدم.",
        "category": "diabetes-care",
        "subcategory": "oral-medications",
        "brand": "DiabCare",
        "manufacturer": "Alexandria Pharma",
        "country_of_origin": "Egypt",
        "price_per_unit": 35.00,
        "currency": "EGP",
        "unit": "box",
        "unit_size": "30 tablets",
        "stock_quantity": 400,
        "min_order_quantity": 1,
        "max_order_quantity": 3,
        "sku": "METF-500-30T",
        "barcode": "6221234567921",
        "hsn_code": "30049099",
        "requires_prescription": True,
        "is_controlled": False,
        "active_ingredient": "Metformin hydrochloride 500mg",
        "dosage_form": "Film-coated tablet",
        "side_effects": "Common: GI upset, diarrhea. Rare: lactic acidosis.",
        "contraindications": "Severe kidney disease, metabolic acidosis",
        "storage_conditions": "Store below 30°C",
        "expiry_date": "2026-09-30",
        "batch_number": "B2024-004",
        "seller_code": "SELLER-001",
        "delivery_method": "standard",
        "delivery_days_min": 1,
        "delivery_days_max": 3,
        "delivery_fee": 30.00,
        "free_delivery_threshold": 200.00,
        "return_policy": "No returns for prescription medications",
        "warranty_period": "N/A",
        "discount_percentage": 5,
        "tax_percentage": 14,
        "product_images": json.dumps([
            "https://example.com/images/metformin-1.jpg"
        ]),
        "specifications": json.dumps({
            "form": "Oval tablet",
            "color": "White",
            "package": "Blister pack"
        }),
        "tags": "diabetes,prescription,blood-sugar,metformin",
        "meta_title": "Metformin 500mg Tablets | Diabetes Management",
        "meta_description": "Metformin 500mg for type 2 diabetes. Prescription required.",
        "is_featured": True,
        "is_active": True,
        "weight_grams": 60,
        "dimensions_cm": "12x10x3"
    },
    {
        "product_code": "MED-005-OMEP",
        "name_en": "Omeprazole 20mg Capsules",
        "name_ar": "أوميبرازول 20 مجم كبسولات",
        "description_en": "Proton pump inhibitor for acid reflux and stomach ulcers. Reduces stomach acid production.",
        "description_ar": "مثبط مضخة البروتون لارتجاع الحمض وقرحة المعدة. يقلل من إنتاج حمض المعدة.",
        "category": "digestive-health",
        "subcategory": "oral-medications",
        "brand": "GastroGuard",
        "manufacturer": "Nile Pharmaceuticals",
        "country_of_origin": "Egypt",
        "price_per_unit": 55.00,
        "currency": "EGP",
        "unit": "box",
        "unit_size": "14 capsules",
        "stock_quantity": 350,
        "min_order_quantity": 1,
        "max_order_quantity": 4,
        "sku": "OMEP-20-14C",
        "barcode": "6221234567938",
        "hsn_code": "30049099",
        "requires_prescription": False,
        "is_controlled": False,
        "active_ingredient": "Omeprazole 20mg",
        "dosage_form": "Enteric-coated capsule",
        "side_effects": "Common: headache, abdominal pain. Long-term: B12 deficiency.",
        "contraindications": "Hypersensitivity to omeprazole",
        "storage_conditions": "Store below 25°C, protect from moisture",
        "expiry_date": "2026-11-30",
        "batch_number": "B2024-005",
        "seller_code": "SELLER-003",
        "delivery_method": "standard",
        "delivery_days_min": 2,
        "delivery_days_max": 4,
        "delivery_fee": 35.00,
        "free_delivery_threshold": 300.00,
        "return_policy": "7 days return for unopened packages",
        "warranty_period": "N/A",
        "discount_percentage": 8,
        "tax_percentage": 14,
        "product_images": json.dumps([
            "https://example.com/images/omeprazole-1.jpg",
            "https://example.com/images/omeprazole-2.jpg"
        ]),
        "specifications": json.dumps({
            "form": "Hard capsule",
            "color": "Pink/White",
            "package": "Blister pack"
        }),
        "tags": "digestive,acid-reflux,ulcer,OTC",
        "meta_title": "Omeprazole 20mg Capsules | Acid Reflux Relief",
        "meta_description": "Omeprazole 20mg for effective acid reflux treatment. Available OTC.",
        "is_featured": False,
        "is_active": True,
        "weight_grams": 35,
        "dimensions_cm": "10x8x2"
    },
    {
        "product_code": "MED-006-LORAT",
        "name_en": "Loratadine 10mg Tablets",
        "name_ar": "لوراتادين 10 مجم أقراص",
        "description_en": "Non-drowsy antihistamine for allergy relief. Treats hay fever, hives, and allergic reactions.",
        "description_ar": "مضاد للهستامين غير مسبب للنعاس لتخفيف الحساسية. يعالج حمى القش والشرى وردود الفعل التحسسية.",
        "category": "allergy-relief",
        "subcategory": "oral-medications",
        "brand": "AllerFree",
        "manufacturer": "Delta Pharma",
        "country_of_origin": "Egypt",
        "price_per_unit": 25.00,
        "currency": "EGP",
        "unit": "box",
        "unit_size": "10 tablets",
        "stock_quantity": 600,
        "min_order_quantity": 1,
        "max_order_quantity": 8,
        "sku": "LORAT-10-10T",
        "barcode": "6221234567945",
        "hsn_code": "30049099",
        "requires_prescription": False,
        "is_controlled": False,
        "active_ingredient": "Loratadine 10mg",
        "dosage_form": "Tablet",
        "side_effects": "Rare: headache, dry mouth, fatigue",
        "contraindications": "Hypersensitivity to loratadine",
        "storage_conditions": "Store at room temperature",
        "expiry_date": "2027-01-31",
        "batch_number": "B2024-006",
        "seller_code": "SELLER-002",
        "delivery_method": "express",
        "delivery_days_min": 1,
        "delivery_days_max": 2,
        "delivery_fee": 40.00,
        "free_delivery_threshold": 150.00,
        "return_policy": "14 days return policy",
        "warranty_period": "N/A",
        "discount_percentage": 12,
        "tax_percentage": 14,
        "product_images": json.dumps([
            "https://example.com/images/loratadine-1.jpg"
        ]),
        "specifications": json.dumps({
            "form": "Round tablet",
            "color": "White",
            "package": "Blister pack"
        }),
        "tags": "allergy,antihistamine,hay-fever,OTC",
        "meta_title": "Loratadine 10mg Non-Drowsy Allergy Relief",
        "meta_description": "Loratadine 10mg tablets for 24-hour allergy relief without drowsiness.",
        "is_featured": True,
        "is_active": True,
        "weight_grams": 25,
        "dimensions_cm": "8x6x2"
    },
    {
        "product_code": "MED-007-SALB",
        "name_en": "Salbutamol Inhaler 100mcg",
        "name_ar": "بخاخ سالبوتامول 100 ميكروجرام",
        "description_en": "Reliever inhaler for asthma and COPD. Provides quick relief from breathing difficulties.",
        "description_ar": "بخاخ مخفف للربو ومرض الانسداد الرئوي المزمن. يوفر راحة سريعة من صعوبات التنفس.",
        "category": "respiratory",
        "subcategory": "inhalers",
        "brand": "BreathEasy",
        "manufacturer": "Cairo Respiratory Solutions",
        "country_of_origin": "Egypt",
        "price_per_unit": 95.00,
        "currency": "EGP",
        "unit": "inhaler",
        "unit_size": "200 doses",
        "stock_quantity": 150,
        "min_order_quantity": 1,
        "max_order_quantity": 2,
        "sku": "SALB-100-200D",
        "barcode": "6221234567952",
        "hsn_code": "30049099",
        "requires_prescription": True,
        "is_controlled": False,
        "active_ingredient": "Salbutamol sulfate 100mcg per dose",
        "dosage_form": "Pressurized inhaler",
        "side_effects": "Common: tremor, headache, increased heart rate",
        "contraindications": "Hypersensitivity to salbutamol",
        "storage_conditions": "Store below 30°C, do not puncture",
        "expiry_date": "2026-08-31",
        "batch_number": "B2024-007",
        "seller_code": "SELLER-001",
        "delivery_method": "express",
        "delivery_days_min": 1,
        "delivery_days_max": 2,
        "delivery_fee": 50.00,
        "free_delivery_threshold": 400.00,
        "return_policy": "No returns for inhalers",
        "warranty_period": "N/A",
        "discount_percentage": 0,
        "tax_percentage": 14,
        "product_images": json.dumps([
            "https://example.com/images/salbutamol-1.jpg",
            "https://example.com/images/salbutamol-2.jpg",
            "https://example.com/images/salbutamol-3.jpg"
        ]),
        "specifications": json.dumps({
            "form": "Aerosol inhaler",
            "color": "Blue",
            "package": "Individual box"
        }),
        "tags": "asthma,respiratory,inhaler,prescription",
        "meta_title": "Salbutamol 100mcg Inhaler | Asthma Relief",
        "meta_description": "Salbutamol inhaler for quick asthma relief. Prescription required.",
        "is_featured": True,
        "is_active": True,
        "weight_grams": 80,
        "dimensions_cm": "15x5x5"
    },
    {
        "product_code": "MED-008-DIAZ",
        "name_en": "Diazepam 5mg Tablets",
        "name_ar": "ديازيبام 5 مجم أقراص",
        "description_en": "Anxiolytic medication for anxiety disorders. Also used for muscle spasms and seizures.",
        "description_ar": "دواء مضاد للقلق لاضطرابات القلق. يستخدم أيضًا لتشنجات العضلات والنوبات.",
        "category": "mental-health",
        "subcategory": "oral-medications",
        "brand": "CalmMed",
        "manufacturer": "Egyptian Neuro Pharma",
        "country_of_origin": "Egypt",
        "price_per_unit": 120.00,
        "currency": "EGP",
        "unit": "box",
        "unit_size": "30 tablets",
        "stock_quantity": 50,
        "min_order_quantity": 1,
        "max_order_quantity": 1,
        "sku": "DIAZ-5-30T",
        "barcode": "6221234567969",
        "hsn_code": "30049099",
        "requires_prescription": True,
        "is_controlled": True,
        "active_ingredient": "Diazepam 5mg",
        "dosage_form": "Tablet",
        "side_effects": "Common: drowsiness, fatigue, muscle weakness. Risk of dependence.",
        "contraindications": "Myasthenia gravis, severe respiratory insufficiency, sleep apnea",
        "storage_conditions": "Store in secure location below 25°C",
        "expiry_date": "2026-05-31",
        "batch_number": "B2024-008",
        "seller_code": "SELLER-001",
        "delivery_method": "special",
        "delivery_days_min": 1,
        "delivery_days_max": 1,
        "delivery_fee": 75.00,
        "free_delivery_threshold": 0,
        "return_policy": "No returns for controlled substances",
        "warranty_period": "N/A",
        "discount_percentage": 0,
        "tax_percentage": 14,
        "product_images": json.dumps([
            "https://example.com/images/diazepam-1.jpg"
        ]),
        "specifications": json.dumps({
            "form": "Round tablet",
            "color": "Yellow",
            "package": "Secure blister pack"
        }),
        "tags": "anxiety,controlled,prescription,mental-health",
        "meta_title": "Diazepam 5mg - Controlled Medication",
        "meta_description": "Diazepam 5mg tablets. Controlled substance, prescription required.",
        "is_featured": False,
        "is_active": True,
        "weight_grams": 40,
        "dimensions_cm": "10x8x2"
    },
    {
        "product_code": "MED-009-IBUP",
        "name_en": "Ibuprofen 400mg Tablets",
        "name_ar": "إيبوبروفين 400 مجم أقراص",
        "description_en": "Non-steroidal anti-inflammatory drug for pain, fever, and inflammation.",
        "description_ar": "دواء مضاد للالتهاب غير ستيرويدي للألم والحمى والالتهاب.",
        "category": "pain-relief",
        "subcategory": "oral-medications",
        "brand": "PainAway",
        "manufacturer": "Giza Pharmaceuticals",
        "country_of_origin": "Egypt",
        "price_per_unit": 22.00,
        "currency": "EGP",
        "unit": "box",
        "unit_size": "20 tablets",
        "stock_quantity": 450,
        "min_order_quantity": 1,
        "max_order_quantity": 6,
        "sku": "IBUP-400-20T",
        "barcode": "6221234567976",
        "hsn_code": "30049099",
        "requires_prescription": False,
        "is_controlled": False,
        "active_ingredient": "Ibuprofen 400mg",
        "dosage_form": "Film-coated tablet",
        "side_effects": "Common: GI upset. Rare: GI bleeding, kidney problems with long-term use.",
        "contraindications": "Active peptic ulcer, severe heart failure, third trimester pregnancy",
        "storage_conditions": "Store below 30°C",
        "expiry_date": "2027-02-28",
        "batch_number": "B2024-009",
        "seller_code": "SELLER-003",
        "delivery_method": "standard",
        "delivery_days_min": 1,
        "delivery_days_max": 3,
        "delivery_fee": 30.00,
        "free_delivery_threshold": 200.00,
        "return_policy": "7 days return for unopened packages",
        "warranty_period": "N/A",
        "discount_percentage": 15,
        "tax_percentage": 14,
        "product_images": json.dumps([
            "https://example.com/images/ibuprofen-1.jpg",
            "https://example.com/images/ibuprofen-2.jpg"
        ]),
        "specifications": json.dumps({
            "form": "Oblong tablet",
            "color": "Pink",
            "package": "Blister pack"
        }),
        "tags": "pain-relief,anti-inflammatory,fever,OTC",
        "meta_title": "Ibuprofen 400mg Tablets | Pain & Inflammation Relief",
        "meta_description": "Ibuprofen 400mg for effective pain and inflammation relief.",
        "is_featured": False,
        "is_active": True,
        "weight_grams": 45,
        "dimensions_cm": "10x8x3"
    },
    {
        "product_code": "MED-010-ZINC",
        "name_en": "Zinc Sulfate 20mg Tablets",
        "name_ar": "كبريتات الزنك 20 مجم أقراص",
        "description_en": "Essential mineral supplement for immune support and wound healing.",
        "description_ar": "مكمل معدني أساسي لدعم المناعة والتئام الجروح.",
        "category": "vitamins",
        "subcategory": "supplements",
        "brand": "ImmunoBoost",
        "manufacturer": "Egyptian Nutrition Labs",
        "country_of_origin": "Egypt",
        "price_per_unit": 35.00,
        "currency": "EGP",
        "unit": "bottle",
        "unit_size": "100 tablets",
        "stock_quantity": 300,
        "min_order_quantity": 1,
        "max_order_quantity": 5,
        "sku": "ZINC-20-100T",
        "barcode": "6221234567983",
        "hsn_code": "21069099",
        "requires_prescription": False,
        "is_controlled": False,
        "active_ingredient": "Zinc sulfate monohydrate equivalent to 20mg elemental zinc",
        "dosage_form": "Tablet",
        "side_effects": "Common: nausea if taken on empty stomach. Rare: metallic taste.",
        "contraindications": "Copper deficiency",
        "storage_conditions": "Store in a cool, dry place",
        "expiry_date": "2027-04-30",
        "batch_number": "B2024-010",
        "seller_code": "SELLER-002",
        "delivery_method": "standard",
        "delivery_days_min": 2,
        "delivery_days_max": 4,
        "delivery_fee": 30.00,
        "free_delivery_threshold": 250.00,
        "return_policy": "30 days return for unopened bottles",
        "warranty_period": "N/A",
        "discount_percentage": 20,
        "tax_percentage": 14,
        "product_images": json.dumps([
            "https://example.com/images/zinc-1.jpg",
            "https://example.com/images/zinc-2.jpg",
            "https://example.com/images/zinc-3.jpg",
            "https://example.com/images/zinc-4.jpg"
        ]),
        "specifications": json.dumps({
            "form": "Round tablet",
            "color": "White",
            "package": "Plastic bottle"
        }),
        "tags": "supplement,immune-support,mineral,zinc,OTC",
        "meta_title": "Zinc Sulfate 20mg - 100 Tablets | Immune Support",
        "meta_description": "Zinc supplement for immune system support and overall health.",
        "is_featured": True,
        "is_active": True,
        "weight_grams": 150,
        "dimensions_cm": "7x7x12"
    }
]

def create_excel_template():
    """Create Excel workbook with products and instructions"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Products sheet
    products_sheet = wb.create_sheet("Products", 0)
    
    # Create DataFrame from products data
    df = pd.DataFrame(products_data)
    
    # Add header row with styling
    header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write headers
    for c_idx, col_name in enumerate(df.columns, 1):
        cell = products_sheet.cell(row=1, column=c_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = products_sheet.cell(row=r_idx, column=c_idx, value=value)
            
            # Apply number format for currency columns
            if df.columns[c_idx-1] in ['price_per_unit', 'delivery_fee', 'free_delivery_threshold']:
                cell.number_format = '#,##0.00 "EGP"'
            
            # Apply percentage format
            elif df.columns[c_idx-1] in ['discount_percentage', 'tax_percentage']:
                cell.number_format = '0"%"'
            
            # Apply date format
            elif df.columns[c_idx-1] == 'expiry_date':
                cell.number_format = 'yyyy-mm-dd'
            
            # Center align specific columns
            if df.columns[c_idx-1] in ['product_code', 'sku', 'barcode', 'requires_prescription', 'is_controlled', 'is_active', 'is_featured']:
                cell.alignment = Alignment(horizontal="center")
    
    # Adjust column widths
    column_widths = {
        'product_code': 15,
        'name_en': 30,
        'name_ar': 30,
        'description_en': 40,
        'description_ar': 40,
        'category': 15,
        'subcategory': 15,
        'brand': 12,
        'manufacturer': 20,
        'country_of_origin': 15,
        'price_per_unit': 12,
        'currency': 8,
        'unit': 8,
        'unit_size': 12,
        'stock_quantity': 12,
        'min_order_quantity': 10,
        'max_order_quantity': 10,
        'sku': 15,
        'barcode': 15,
        'hsn_code': 12,
        'requires_prescription': 12,
        'is_controlled': 10,
        'active_ingredient': 25,
        'dosage_form': 15,
        'side_effects': 30,
        'contraindications': 25,
        'storage_conditions': 20,
        'expiry_date': 12,
        'batch_number': 12,
        'seller_code': 12,
        'delivery_method': 12,
        'delivery_days_min': 10,
        'delivery_days_max': 10,
        'delivery_fee': 12,
        'free_delivery_threshold': 15,
        'return_policy': 25,
        'warranty_period': 12,
        'discount_percentage': 10,
        'tax_percentage': 10,
        'product_images': 40,
        'specifications': 30,
        'tags': 20,
        'meta_title': 30,
        'meta_description': 35,
        'is_featured': 10,
        'is_active': 10,
        'weight_grams': 10,
        'dimensions_cm': 12
    }
    
    for idx, col_name in enumerate(df.columns, 1):
        col_letter = products_sheet.cell(row=1, column=idx).column_letter
        products_sheet.column_dimensions[col_letter].width = column_widths.get(col_name, 15)
    
    # Freeze the header row
    products_sheet.freeze_panes = 'A2'
    
    # Add filters
    products_sheet.auto_filter.ref = products_sheet.dimensions
    
    # Create Instructions sheet
    instructions_sheet = wb.create_sheet("Instructions", 1)
    
    instructions = [
        ["Medicum Egypt - Product Import Template Instructions", ""],
        ["", ""],
        ["Overview", "This Excel template is designed for bulk importing products into the Medicum Egypt medical e-commerce platform."],
        ["", ""],
        ["Required Fields", "All fields marked as required must be filled for successful import:"],
        ["", "• product_code: Unique identifier for each product"],
        ["", "• name_en: Product name in English"],
        ["", "• name_ar: Product name in Arabic"],
        ["", "• price_per_unit: Price in Egyptian Pounds (EGP)"],
        ["", "• seller_code: Must match an existing seller in the system"],
        ["", ""],
        ["Field Descriptions", ""],
        ["product_code", "Unique product identifier (e.g., MED-001-PARA)"],
        ["name_en/name_ar", "Product name in English and Arabic"],
        ["description_en/description_ar", "Detailed product description in both languages"],
        ["category", "Main category (e.g., pain-relief, antibiotics, vitamins)"],
        ["subcategory", "Subcategory (e.g., oral-medications, supplements, inhalers)"],
        ["brand", "Brand name"],
        ["manufacturer", "Manufacturer company name"],
        ["country_of_origin", "Country where product is manufactured"],
        ["price_per_unit", "Price per unit in EGP"],
        ["currency", "Currency code (default: EGP)"],
        ["unit", "Unit of sale (e.g., box, bottle, inhaler)"],
        ["unit_size", "Size/quantity in unit (e.g., 20 tablets, 60 softgels)"],
        ["stock_quantity", "Current stock quantity"],
        ["min_order_quantity", "Minimum order quantity"],
        ["max_order_quantity", "Maximum order quantity"],
        ["sku", "Stock Keeping Unit"],
        ["barcode", "Product barcode"],
        ["hsn_code", "Harmonized System Nomenclature code"],
        ["requires_prescription", "TRUE/FALSE - Whether prescription is required"],
        ["is_controlled", "TRUE/FALSE - Whether it's a controlled substance"],
        ["active_ingredient", "Active pharmaceutical ingredient with strength"],
        ["dosage_form", "Form of medication (tablet, capsule, syrup, etc.)"],
        ["side_effects", "Potential side effects"],
        ["contraindications", "Medical conditions where product shouldn't be used"],
        ["storage_conditions", "Storage requirements"],
        ["expiry_date", "Product expiry date (YYYY-MM-DD format)"],
        ["batch_number", "Manufacturing batch number"],
        ["seller_code", "Seller identifier (must exist in system)"],
        ["delivery_method", "standard/express/special"],
        ["delivery_days_min/max", "Minimum and maximum delivery days"],
        ["delivery_fee", "Delivery charge in EGP"],
        ["free_delivery_threshold", "Order amount for free delivery"],
        ["return_policy", "Return policy description"],
        ["warranty_period", "Warranty period if applicable"],
        ["discount_percentage", "Discount percentage (0-100)"],
        ["tax_percentage", "Tax percentage (typically 14% in Egypt)"],
        ["product_images", "JSON array of image URLs"],
        ["specifications", "JSON object with additional specifications"],
        ["tags", "Comma-separated tags for search"],
        ["meta_title", "SEO meta title"],
        ["meta_description", "SEO meta description"],
        ["is_featured", "TRUE/FALSE - Whether to feature product"],
        ["is_active", "TRUE/FALSE - Whether product is active"],
        ["weight_grams", "Product weight in grams"],
        ["dimensions_cm", "Product dimensions (LxWxH in cm)"],
        ["", ""],
        ["Data Validation Rules", ""],
        ["", "• product_code must be unique"],
        ["", "• prices must be positive numbers"],
        ["", "• dates must be in YYYY-MM-DD format"],
        ["", "• boolean fields must be TRUE or FALSE"],
        ["", "• JSON fields must be valid JSON format"],
        ["", "• seller_code must match existing seller"],
        ["", ""],
        ["Import Process", ""],
        ["", "1. Fill in all required fields for each product"],
        ["", "2. Validate data according to the rules above"],
        ["", "3. Save the file in Excel format (.xlsx)"],
        ["", "4. Use the admin panel's import feature to upload"],
        ["", "5. Review any validation errors and correct them"],
        ["", "6. Confirm import to add products to database"],
        ["", ""],
        ["Sample Sellers", ""],
        ["SELLER-001", "PharmaCare Egypt - Main pharmaceutical supplier"],
        ["SELLER-002", "VitaHealth - Vitamins and supplements specialist"],
        ["SELLER-003", "MediCore - General medical supplies"],
        ["", ""],
        ["Categories Available", ""],
        ["", "• pain-relief"],
        ["", "• antibiotics"],
        ["", "• vitamins"],
        ["", "• diabetes-care"],
        ["", "• digestive-health"],
        ["", "• allergy-relief"],
        ["", "• respiratory"],
        ["", "• mental-health"],
        ["", "• first-aid"],
        ["", "• personal-care"],
        ["", ""],
        ["Support", ""],
        ["", "For assistance with product import, contact: support@medicumegypt.com"],
        ["", "Generated on: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    ]
    
    # Style for instructions sheet
    title_font = Font(bold=True, size=14, color="1E88E5")
    header_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = instructions_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Apply formatting
            if row_idx == 1:
                cell.font = title_font
            elif row_data[0] in ["Overview", "Required Fields", "Field Descriptions", "Data Validation Rules", "Import Process", "Sample Sellers", "Categories Available", "Support"]:
                cell.font = header_font
                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            else:
                cell.font = normal_font
    
    # Set column widths for instructions
    instructions_sheet.column_dimensions['A'].width = 25
    instructions_sheet.column_dimensions['B'].width = 80
    
    # Create Validation Lists sheet
    validation_sheet = wb.create_sheet("Validation Lists", 2)
    
    validation_data = {
        "Categories": ["pain-relief", "antibiotics", "vitamins", "diabetes-care", "digestive-health", 
                      "allergy-relief", "respiratory", "mental-health", "first-aid", "personal-care"],
        "Subcategories": ["oral-medications", "supplements", "inhalers", "topical", "injections", "syrups"],
        "Delivery Methods": ["standard", "express", "special"],
        "Seller Codes": ["SELLER-001", "SELLER-002", "SELLER-003"],
        "Currency": ["EGP", "USD"],
        "Units": ["box", "bottle", "inhaler", "tube", "pack", "vial"],
        "Dosage Forms": ["Tablet", "Capsule", "Syrup", "Injection", "Cream", "Ointment", "Inhaler", "Drops", "Gel", "Spray"]
    }
    
    # Write validation lists
    col_idx = 1
    for list_name, list_values in validation_data.items():
        # Header
        cell = validation_sheet.cell(row=1, column=col_idx, value=list_name)
        cell.font = header_font
        cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        
        # Values
        for row_idx, value in enumerate(list_values, 2):
            validation_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        col_idx += 1
    
    # Adjust column widths
    for col in range(1, len(validation_data) + 1):
        validation_sheet.column_dimensions[validation_sheet.cell(row=1, column=col).column_letter].width = 20
    
    # Add data validation to Products sheet
    # Category validation
    category_validation = DataValidation(
        type="list",
        formula1="='Validation Lists'.A2:A11",
        allow_blank=True
    )
    category_validation.error = "Please select a valid category from the list"
    category_validation.errorTitle = "Invalid Category"
    products_sheet.add_data_validation(category_validation)
    category_validation.add("F2:F1000")  # Apply to category column
    
    # Delivery method validation
    delivery_validation = DataValidation(
        type="list",
        formula1="='Validation Lists'.C2:C4",
        allow_blank=True
    )
    products_sheet.add_data_validation(delivery_validation)
    delivery_validation.add("AE2:AE1000")  # Apply to delivery_method column
    
    # Boolean validation for prescription and controlled fields
    bool_validation = DataValidation(
        type="list",
        formula1='"TRUE,FALSE"',
        allow_blank=False
    )
    products_sheet.add_data_validation(bool_validation)
    bool_validation.add("U2:U1000")  # requires_prescription
    bool_validation.add("V2:V1000")  # is_controlled
    bool_validation.add("AV2:AV1000")  # is_featured
    bool_validation.add("AW2:AW1000")  # is_active
    
    # Save the workbook
    output_file = "/home/user/webapp/medicum_egypt_products_import_template.xlsx"
    wb.save(output_file)
    print(f"✅ Excel template created successfully: {output_file}")
    print(f"📊 Total products: {len(products_data)}")
    print(f"📑 Sheets created: Products ({len(products_data)} items), Instructions, Validation Lists")
    
    return output_file

if __name__ == "__main__":
    create_excel_template()